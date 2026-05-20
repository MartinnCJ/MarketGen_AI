"""Reports router — aggregated KPIs and export."""
from __future__ import annotations

import io
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse

from app.dependencies.auth import CurrentUser, get_current_user
from app.services.firestore_service import (
    books_repo, proposals_repo, customers_repo,
)

router = APIRouter(prefix="/reports", tags=["Reports"])


@router.get("/overview")
async def reports_overview(user: CurrentUser = Depends(get_current_user)):
    """Return high-level KPIs for the current user."""
    # Run counts in parallel (Firestore async)
    import asyncio
    total_books, total_proposals, total_customers = await asyncio.gather(
        books_repo.count(filters=[("userId", "==", user.sub)]),
        proposals_repo.count(filters=[("userId", "==", user.sub)]),
        customers_repo.count(filters=[("userId", "==", user.sub)]),
    )

    # Books — chapters count via listing
    books = await books_repo.list(
        filters=[("userId", "==", user.sub)], limit=200,
    )
    chapters_generated = sum(b.get("chapterCount") or 0 for b in books)

    # Status breakdown
    status_map: dict[str, int] = {}
    for b in books:
        status_map[b.get("status", "draft")] = status_map.get(b.get("status", "draft"), 0) + 1

    books_by_status = [{"name": k, "value": v} for k, v in status_map.items()]

    return {
        "kpis": {
            "totalBooks":       total_books,
            "chaptersGenerated": chapters_generated,
            "proposalsSent":    total_proposals,
            "totalCustomers":   total_customers,
        },
        "booksByStatus": books_by_status,
        "generatedAt":   datetime.now(timezone.utc).isoformat(),
    }

@router.get("/dashboard")
async def reports_dashboard(user: CurrentUser = Depends(get_current_user)):
    """Alias for dashboard KPI endpoint expected by Swagger/frontend."""
    return await reports_overview(user)

@router.get("/books")
async def reports_books(
    limit:  int = Query(50, ge=1, le=200),
    user:   CurrentUser = Depends(get_current_user),
):
    """Return per-book stats."""
    books = await books_repo.list(
        filters=[("userId", "==", user.sub)],
        order_by="updatedAt", order_direction="DESCENDING",
        limit=limit,
    )
    return {"items": books, "total": len(books)}

@router.get("/proposals")
async def reports_proposals(
    limit: int = Query(10, ge=1, le=100),
    user: CurrentUser = Depends(get_current_user),
):
    proposals = await proposals_repo.list(
        filters=[("userId", "==", user.sub)],
        order_by="createdAt",
        order_direction="DESCENDING",
        limit=500,
    )

    status_map = {}

    for proposal in proposals:
        status = proposal.get("status", "draft")
        amount = proposal.get("totalAmount") or proposal.get("total") or 0

        if status not in status_map:
            status_map[status] = {"label": status, "count": 0, "totalValue": 0}

        status_map[status]["count"] += 1
        status_map[status]["totalValue"] += float(amount or 0)

    return {
        "dimension": "status",
        "data": list(status_map.values())[:limit],
        "timeSeries": [],
    }

@router.get("/content")
async def reports_content(
    limit: int = Query(50, ge=1, le=200),
    user: CurrentUser = Depends(get_current_user),
):
    books = await books_repo.list(
        filters=[("userId", "==", user.sub)],
        order_by="updatedAt",
        order_direction="DESCENDING",
        limit=limit,
    )

    return {
        "data": [
            {
                "bookId": book.get("id"),
                "bookTitle": book.get("title", ""),
                "status": book.get("status", "draft"),
                "assetCount": book.get("assetCount", 0),
                "downloadCount": book.get("downloadCount", 0),
                "createdAt": str(book.get("createdAt", "")),
            }
            for book in books
        ]
    }

@router.get("/export")
async def export_report(
    format: str = Query("xlsx", pattern="^(xlsx|csv)$"),
    user:   CurrentUser = Depends(get_current_user),
):
    """Export a simple overview report as Excel or CSV."""
    books = await books_repo.list(
        filters=[("userId", "==", user.sub)], limit=500,
    )

    rows = [
        {
            "Título":     b.get("title", ""),
            "Estado":     b.get("status", ""),
            "Tipo":       b.get("contentType", ""),
            "Capítulos":  b.get("chapterCount", 0),
            "Actualizado": str(b.get("updatedAt", "")),
        }
        for b in books
    ]

    if format == "csv":
        import csv
        buf = io.StringIO()
        if rows:
            writer = csv.DictWriter(buf, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
        return StreamingResponse(
            io.BytesIO(buf.getvalue().encode()),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=report.csv"},
        )

    # xlsx via openpyxl
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return {"error": "openpyxl not installed"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Libros"

    # Header
    headers = ["Título", "Estado", "Tipo", "Capítulos", "Actualizado"]
    header_fill = PatternFill("solid", fgColor="6366F1")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 25

    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row[key])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=report.xlsx"},
    )
